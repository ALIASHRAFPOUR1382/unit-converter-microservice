from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import crud
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["export"])


def create_excel_file(todos: list, conversions: list) -> str:
    """
    Create Excel file with todos and conversion history
    Returns the file path
    """
    wb = Workbook()
    
    # Remove default sheet if it exists
    if len(wb.sheetnames) > 0:
        wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Always create Todos sheet
    todos_sheet = wb.create_sheet("Todos", 0)
    
    # Headers for Todos
    todos_headers = ["ID", "Title", "Description", "Completed", "Created At", "Updated At"]
    todos_sheet.append(todos_headers)
    
    # Style headers
    for cell in todos_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add todos data
    logger.info(f"Adding {len(todos)} todos to Excel sheet")
    for todo in todos:
        todos_sheet.append([
            todo.id,
            todo.title,
            todo.description or "",
            "Yes" if todo.completed else "No",
            todo.created_at.strftime("%Y-%m-%d %H:%M:%S") if todo.created_at else "",
            todo.updated_at.strftime("%Y-%m-%d %H:%M:%S") if todo.updated_at else ""
        ])
    
    # Auto-adjust column widths for Todos
    for column in todos_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        todos_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Always create Conversion History sheet
    conv_sheet = wb.create_sheet("Conversion History", 1)
    
    # Headers for Conversion History
    conv_headers = ["ID", "Value", "From Unit", "To Unit", "Result", "Unit Type", "Created At"]
    conv_sheet.append(conv_headers)
    
    # Style headers
    for cell in conv_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add conversion data
    logger.info(f"Adding {len(conversions)} conversions to Excel sheet")
    print(f"[EXCEL] Adding {len(conversions)} conversions to Excel sheet")
    
    if len(conversions) == 0:
        print("[EXCEL] WARNING: No conversions to add to Excel!")
    
    for idx, conv in enumerate(conversions):
        try:
            conv_sheet.append([
                conv.id,
                conv.value,
                conv.from_unit,
                conv.to_unit,
                conv.result,
                conv.unit_type,
                conv.created_at.strftime("%Y-%m-%d %H:%M:%S") if conv.created_at else ""
            ])
            if idx < 3:  # Print first 3 for debugging
                print(f"[EXCEL] Added conversion {idx+1}: ID={conv.id}, Value={conv.value}, From={conv.from_unit}, To={conv.to_unit}")
        except Exception as e:
            print(f"[EXCEL] Error adding conversion {idx+1}: {str(e)}")
            logger.error(f"Error adding conversion to Excel: {str(e)}")
    
    print(f"[EXCEL] Total rows in Conversion History sheet: {conv_sheet.max_row} (should be {len(conversions) + 1} including header)")
    
    # Verify data was written
    if conv_sheet.max_row != len(conversions) + 1:
        print(f"[EXCEL] ERROR: Row count mismatch! Expected {len(conversions) + 1} rows, got {conv_sheet.max_row}")
        logger.error(f"Row count mismatch in Conversion History sheet: expected {len(conversions) + 1}, got {conv_sheet.max_row}")
    
    # Auto-adjust column widths for Conversion History
    for column in conv_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        conv_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    BASE_DIR = Path(__file__).resolve().parent.parent.parent
    exports_dir = BASE_DIR / "exports"
    exports_dir.mkdir(exist_ok=True)
    
    filename = f"database_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = exports_dir / filename
    
    wb.save(filepath)
    
    return str(filepath)


@router.get("/excel")
def export_to_excel(db: Session = Depends(get_db)):
    """
    Export all database data (Todos and Conversion History) to Excel file
    
    Returns an Excel file with two sheets:
    - Todos: All todo items
    - Conversion History: All conversion records
    """
    try:
        # Direct database query to verify data exists
        from app import models
        direct_count = db.query(models.ConversionHistory).count()
        print(f"[EXPORT] Direct DB query count: {direct_count} conversions")
        logger.info(f"Direct DB query found {direct_count} conversions")
        
        # Get all todos without pagination
        todos = crud.get_all_todos(db=db)
        logger.info(f"Exporting {len(todos)} todos")
        print(f"[EXPORT] Found {len(todos)} todos in database")
        
        # Get all conversion history without pagination - try multiple methods
        conversions = crud.get_all_conversion_history(db=db)
        logger.info(f"Exporting {len(conversions)} conversions")
        print(f"[EXPORT] Found {len(conversions)} conversions using crud.get_all_conversion_history")
        
        # If no conversions found, try direct query
        if len(conversions) == 0 and direct_count > 0:
            print("[EXPORT] WARNING: crud returned 0 but direct query found data! Using direct query...")
            conversions = db.query(models.ConversionHistory).order_by(models.ConversionHistory.created_at.desc()).all()
            print(f"[EXPORT] Direct query returned {len(conversions)} conversions")
        
        # Debug: Print first few conversions if any
        if conversions:
            print(f"[EXPORT] First conversion: ID={conversions[0].id}, Value={conversions[0].value}, From={conversions[0].from_unit}, To={conversions[0].to_unit}")
            print(f"[EXPORT] Last conversion: ID={conversions[-1].id}, Value={conversions[-1].value}, From={conversions[-1].from_unit}, To={conversions[-1].to_unit}")
        else:
            print("[EXPORT] WARNING: No conversions found in database!")
            print(f"[EXPORT] Direct count was: {direct_count}")
        
        # Create Excel file
        filepath = create_excel_file(todos, conversions)
        logger.info(f"Excel file created at: {filepath} with {len(todos)} todos and {len(conversions)} conversions")
        print(f"[EXPORT] Excel file created: {filepath} with {len(todos)} todos and {len(conversions)} conversions")
        
        return FileResponse(
            path=filepath,
            filename=os.path.basename(filepath),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")


@router.get("/excel/todos")
def export_todos_to_excel(db: Session = Depends(get_db)):
    """
    Export only Todos to Excel file
    """
    try:
        todos = crud.get_all_todos(db=db)
        logger.info(f"Exporting {len(todos)} todos only")
        filepath = create_excel_file(todos, [])
        
        return FileResponse(
            path=filepath,
            filename=os.path.basename(filepath),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")


@router.get("/excel/conversions")
def export_conversions_to_excel(db: Session = Depends(get_db)):
    """
    Export only Conversion History to Excel file
    """
    try:
        conversions = crud.get_all_conversion_history(db=db)
        logger.info(f"Exporting {len(conversions)} conversions only")
        print(f"[EXPORT] Found {len(conversions)} conversions for export")
        filepath = create_excel_file([], conversions)
        
        return FileResponse(
            path=filepath,
            filename=os.path.basename(filepath),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")


@router.get("/debug/counts")
def get_data_counts(db: Session = Depends(get_db)):
    """
    Debug endpoint to check how many records are in the database
    """
    try:
        from app import models
        from sqlalchemy import desc
        
        # Direct database queries
        direct_todos_count = db.query(models.Todo).count()
        direct_conversions_count = db.query(models.ConversionHistory).count()
        
        # Using CRUD functions
        todos = crud.get_all_todos(db=db)
        conversions = crud.get_all_conversion_history(db=db)
        
        # Get all conversions directly
        all_conversions_direct = db.query(models.ConversionHistory).order_by(desc(models.ConversionHistory.created_at)).all()
        
        return {
            "direct_db_counts": {
                "todos": direct_todos_count,
                "conversions": direct_conversions_count
            },
            "crud_counts": {
                "todos": len(todos),
                "conversions": len(conversions)
            },
            "direct_query_count": len(all_conversions_direct),
            "todos": [{"id": t.id, "title": t.title} for t in todos[:5]],  # First 5
            "conversions_crud": [
                {
                    "id": c.id,
                    "value": c.value,
                    "from_unit": c.from_unit,
                    "to_unit": c.to_unit,
                    "result": c.result,
                    "unit_type": c.unit_type,
                    "created_at": str(c.created_at) if c.created_at else None
                } for c in conversions[:10]  # First 10
            ],
            "conversions_direct": [
                {
                    "id": c.id,
                    "value": c.value,
                    "from_unit": c.from_unit,
                    "to_unit": c.to_unit,
                    "result": c.result,
                    "unit_type": c.unit_type,
                    "created_at": str(c.created_at) if c.created_at else None
                } for c in all_conversions_direct[:10]  # First 10
            ]
        }
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=f"Error getting counts: {str(e)}\n{traceback.format_exc()}")
